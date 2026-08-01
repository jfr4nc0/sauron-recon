from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

from sauron_recon.application.ports import SearchResult
from sauron_recon.domain.changes import ListingChange


def _clean(text: str) -> str:
    text = re.sub(r"!\[[^\]]*\]\([^)]*\)", "", text)
    text = re.sub(r"https?://\S+", "", text)
    text = re.sub(r"\s{3,}", "  ", text).strip()
    return text


def render_xlsx(result: SearchResult, changes: Iterable[ListingChange], output: str | Path) -> Path:
    """Render an Excel spreadsheet with all observed listings."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("XLSX reports require the 'openpyxl' package") from exc

    change_map: dict[str, ListingChange] = {}
    for change in changes:
        change_map.setdefault(change.listing.identity, change)

    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    if wb.active is None:
        wb.create_sheet()

    # --- Sheet 1: Resumen ---
    ws1 = wb.active
    ws1.title = "Resumen"
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="243447")
    normal_font = Font(name="Calibri", size=11)

    ws1["A1"] = "Sauron Recon"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws1["A2"] = f"Corrida: {result.run_id}"
    ws1["A3"] = f"Fecha: {result.started_at.strftime('%Y-%m-%d %H:%M:%S UTC')}"

    new_count = sum(1 for c in change_map.values() if c.kind == "new")
    changed_count = sum(1 for c in change_map.values() if c.kind == "changed")

    ws1["A5"] = "Listings observados"
    ws1["B5"] = len(result.listings)
    ws1["A6"] = "Nuevos"
    ws1["B6"] = new_count
    ws1["A7"] = "Modificados"
    ws1["B7"] = changed_count
    ws1["A8"] = "Duplicados cross-source"
    ws1["B8"] = len(result.duplicate_candidates)
    ws1["A9"] = "Fuentes con error"
    ws1["B9"] = len(result.failures)
    for row in range(5, 10):
        ws1[f"A{row}"].font = normal_font
        ws1[f"B{row}"].font = normal_font

    if result.failures:
        ws1["A11"] = "Advertencias"
        ws1["A11"].font = Font(bold=True)
        for i, failure in enumerate(result.failures):
            ws1[f"A{12 + i}"] = f"[{failure.source}] {failure.error_type}: {failure.message}"
            ws1[f"A{12 + i}"].font = Font(size=9, italic=True)

    # --- Sheet 2: Avisos ---
    ws2 = wb.create_sheet("Avisos")
    headers = ["Estado", "Fuente", "Aviso", "Precio", "Moneda", "Superficie (m2)", "Expensas", "Zona", "Disponibilidad", "Enlace"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for i, listing in enumerate(result.listings, 2):
        change = change_map.get(listing.identity)
        kind = change.kind if change else "observado"
        ws2.cell(row=i, column=1, value=kind)
        ws2.cell(row=i, column=2, value=listing.source)
        ws2.cell(row=i, column=3, value=listing.title)
        ws2.cell(row=i, column=4, value=float(listing.price) if listing.price is not None else None)
        ws2.cell(row=i, column=5, value=listing.currency)
        ws2.cell(row=i, column=6, value=float(listing.area_m2) if listing.area_m2 is not None else None)
        ws2.cell(row=i, column=7, value=float(listing.expenses) if listing.expenses is not None else None)
        ws2.cell(row=i, column=8, value=listing.address)
        ws2.cell(row=i, column=9, value=listing.availability)
        ws2.cell(row=i, column=10, value=listing.url)

    # Column widths
    widths = [12, 14, 50, 14, 10, 16, 14, 30, 16, 60]
    for col, width in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws2.freeze_panes = "A2"

    # Auto-filter on all data
    if len(result.listings) > 0:
        ws2.auto_filter.ref = f"A1:J{i}"

    # Alternating row colors
    alt_fill = PatternFill("solid", fgColor="EEF3F7")
    for row in range(2, len(result.listings) + 2):
        if row % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws2.cell(row=row, column=col).fill = alt_fill

    # Number formats
    for row in range(2, len(result.listings) + 2):
        ws2.cell(row=row, column=4).number_format = "#,##0"
        ws2.cell(row=row, column=6).number_format = "#,##0"
        ws2.cell(row=row, column=7).number_format = "#,##0"

    wb.save(str(output_path))
    return output_path
